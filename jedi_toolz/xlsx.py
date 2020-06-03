# AUTOGENERATED! DO NOT EDIT! File to edit: xlsx.ipynb (unless otherwise specified).

# __all__ = "to_xlsx ColumnFormat format_sheet append_sheet".split()

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
from pathlib import Path
import toolz.curried as tz
from fnmatch import fnmatch
from functools import partial
from typing import (NewType, Dict, List, Iterable, Any, Optional, Union,
    Tuple, NamedTuple, Literal, Pattern)
from pydantic import (BaseModel, confloat, conint, FilePath, validator,
    ValidationError, Field, validate_arguments)
from pydantic.color import Color
from enum import Enum
from collections import Counter


@validate_arguments
def get_wb(file_path: FilePath, read_only: bool=True) -> Workbook:
    return openpyxl.load_workbook(str(file_path), read_only=read_only)


@validate_arguments
def get_sheet_names(file_path: FilePath, *glob:str):
    sheets = get_wb(file_path).sheetnames
    globs = ["*"] if len(glob) == 0 else glob
    return [
        s for s in sheets
        for g in globs
        if fnmatch(s, g)
    ]


class Format(BaseModel):
    glob: str
    number_format: str="General"
    width: float=8.7
    align: Literal["left", "center", "right", "fill", "justify",
        "center_across", "distributed"]="left"
    valign: Literal["top", "vcenter","bottom", "vjustify",
        "vdistributed"]="top"
    text_wrap: bool=True
    bold: bool=False
    font_name: str="Calibri"
    font_size: float=11
    font_color: Color


class Col(BaseModel):
    name: str=Field(description="Column Name|first cell in the column")
    letter: str=Field(
        regex="([A-Z])|([A-Z][A-Z])|([A-X][A-F][A-D])",
        description="Column letter")
    index: int=Field(ge=1, le=16384, description="Index of Column 1-based")
    data_type: Literal["date", "text", "number"]=Field(
        description="Identifier for how the data is stored in Excel")


class WS(BaseModel):
    name: str=Field(min_length=1, max_length=31)
    cols: List[Col]

    @classmethod
    def from_file_path(cls, file_path: FilePath, sheet_name: str, *, row_limit: int=100):
        """Help function to populate the columns of a sheet."""
        wb = get_wb(file_path)
        ws = wb[sheet_name]
        rows = tz.take(row_limit, ws.rows)
        header = next(rows)
        names = [c.value for c in header]
        letters = [c.column_letter for c in header]
        indices = [c.column for c in header]
        data_types = tz.pipe(
            rows
            # For each row, create a dict usng names has keys
            ,tz.map(lambda row: dict(zip(names, row)))
            # Get the .xlsx data_type for each cell
            ,tz.map(tz.valmap(lambda cell: cell.data_type))
            # Combine cells into a list per column
            ,tz.merge_with(list)
            # Count the cells for each data type in the column
            ,tz.valmap(tz.frequencies)
            # Consolidate types
            ,tz.valmap(lambda freq: (
                # If at least 1 "d"
                "date" if "d" in freq else
                # If at least 1 "s"
                "text" if "s" in freq else
                # If at least 1 "n"
                "number" if "n" in freq else
                str(freq)
            ))
            ,lambda d: [v for k, v in d.items()]
        )

        cols = [
            Col(name=N, letter=L, index=I, data_type=D)
            for N, L, I, D in zip(names, letters, indices, data_types)
        ]
        return cls(name=sheet_name, cols=cols)


@validate_arguments
def xlsx(file_path: FilePath, *glob: str) -> List[WS]:
    return tz.pipe(
        get_sheet_names(file_path, *glob)
        ,tz.map(partial(WS.from_file_path, file_path))
        ,list
    )


from pprint import pprint
file_path = Path("~/py-analysis/appraisal-fees/appraisal_fees/reports/Fee Reports.xlsx").expanduser()
# pprint(sheet_names(file_path, "Region*", "State*"))
reports = xlsx(file_path, "Order-Summary", "State-County")
for sheet in reports:
    print(f"{sheet.name=}")
    for col in sheet.cols:
        print(f"  {col}")

@validate_arguments
def format_sheet(
    file_path: FilePath,
    sheet_name: str,
    *format: Format,
    table_name: str="",
    freeze_panes: "cell_refrence"="A2"
    ) -> Path:
    """Formats an existing sheet and optionally adds the sheet as a
    Worksheet table with a default table format.

    Provided formats for the 'patterns' argument use a Unix shell-style
    wildcard as the key and the value is an XLSXFormat. The XLSXFormat is
    typically created by using the column_format function.

    Unix shell-style pattern
    ╒═══════════╤══════════════════════════════════╕
    │ Pattern   │ Meaning                          │
    ╞═══════════╪══════════════════════════════════╡
    │ *         │ matches everything               │
    ├───────────┼──────────────────────────────────┤
    │ ?         │ matches any single character     │
    ├───────────┼──────────────────────────────────┤
    │ [seq]     │ matches any character in seq     │
    ├───────────┼──────────────────────────────────┤
    │ [!seq]    │ matches any character not in seq │
    ╘═══════════╧══════════════════════════════════╛
    """
    # Load file
    wb = get_wb(file_path, read_only=False)
    info = WS.from_file_path(file_path, sheet_name)
    ws = wb[sheet_name]

    for col in info.cols:
        cells = ws[col.letter]
        header = cells[0]
        data = cells[1:]

        # Set width first so that text wrap works
        ws.column_dimensions[letter].width = fmt.width

        # Format header cell
        header.alignment = fmt.header.alignment
        header.number_format = fmt.header.number_format
        header.font = fmt.header.font

        # Add alignment and number format
        for cell in cells:
            cell.alignment = fmt.cell.alignment
            cell.number_format = fmt.cell.number_format
            cell.font = fmt.cell.font

    # Freeze panes
    ws.freeze_panes = ws[freeze_panes]

    # Add table
    table_name = table_name or sheet_name
    table_name = table_name.replace(" ", "_")
    table = Table(displayName=table_name, ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showColumnStripes=False, showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save file
    wb.save(str(file_path))
    wb.close()

    return file_path

# print(WS.from_file(file_path, "Order-Summary"))

# class XLSX(BaseModel):
#     file_path: FilePath
#     sheets: List[Sheet]

#     @classmethod
#     def from_file(cls, file_path: FilePath, sheet_names: SheetNames=None,
#         **kwargs):
#         wb = WB.from_file(file_path, **kwargs)

#     def file_path(self) -> FilePath:
#         return self.wb.file_path

#     def __str__(self):
#         indent = "  "

#         # XLSX header
#         text = f"{self.__class__.__name__}("
#         text += f"\n{indent}file_path = '{self.file_path}',"

#         # sheets header
#         text += f"\n{indent}sheets = ["

#         for sheet in self.sheets:

#             # Sheet
#             text += f"\n{indent*2}{sheet.__class__.__name__}("
#             text += f"\n{indent*3}sheet_name = '{sheet.sheet_name}',"
#             text += f"\n{indent*3}columns = ["

#             # Column
#             for column in sheet.columns:
#                 text += f"\n{indent*4}{column.__class__.__name__}"
#                 text += f"({column}),"

#             # Sheet footer
#             text += f"\n{indent*2}),"

#         # sheets footer
#         text += f"\n{indent}],"

#         # XLSX footer
#         text += "\n)"
#         return text


# Row = List[Cell]
# def get_columns(rows: Iterable[Row]) -> List[Column]:
#     """Help function to populate the columns of a sheet."""
#     rows = iter(rows)
#     header = next(rows)
#     names = [c.value for c in header]
#     letters = [c.column_letter for c in header]
#     indices = [c.column for c in header]
#     data_types = tz.pipe(
#         rows
#         # For each row, create a dict usng names has keys
#         ,tz.map(lambda row: dict(zip(names, row)))
#         # Get the .xlsx data_type for each cell
#         ,tz.map(tz.valmap(lambda cell: cell.data_type))
#         # Combine cells into a list per column
#         ,tz.merge_with(list)
#         # Count the cells for each data type in the column
#         ,tz.valmap(tz.frequencies)
#         # Consolidate types
#         ,tz.valmap(lambda freq: (
#             # If at least 1 "d"
#             "date" if "d" in freq else
#             # If at least 1 "s"
#             "text" if "s" in freq else
#             # If at least 1 "n"
#             "number" if "n" in freq else
#             str(freq)
#         ))
#         ,lambda d: [v for k, v in d.items()]
#     )

#     return [
#         Column(column_name=N, letter=L, index=I, data_type=D)
#         for N, L, I, D in zip(names, letters, indices, data_types)
#     ]



# def listify(arg: Any) -> List[Any]:
#     if arg is None: return []
#     if isinstance(arg, (dict, str): return [arg]
#     if isinstance(arg, (tuple, list)): return list(arg)
#     else: return [arg]

# WorkbookPath = Union[FilePath, Workbook]

# def get_workbook(wb_or_path: WorkbookPath, **kwargs) -> Workbook:
#     if isinstance(wb_or_path, Workbook):
#         return wb_or_path
#     else:
#         return openpyxl.load_workbook(str(wb_or_path), **kwargs)


# SheetNames = Union[str,List[str]]

# @validate_arguments
# def get_sheets(wb_or_path: WorkbookPath, *, sheet_names: SheetNames=None,
#     row_limit: int=100, wb: Workbook=None) -> List[Sheet]:
#     """Helper method to populate Sheet and Column types using openpyxl."""

#     wb = wb or openpyxl.load_workbook(str(file_path), read_only=True)
#     sheet_names = listify(sheet_names) if sheet_names else wb.sheetnames

#     sheets = []

#     # Get attrs for each sheet name in the workbook
#     for sheet_name in sheet_names:

#         # openpyxl worksheet reference
#         ws = wb[sheet_name]

#         # Retrieving data for sheet only works lazily using rows
#         rows = tz.take(row_limit, ws.rows)
#         columns = get_columns(rows)
#         sheet = Sheet(sheet_name=sheet_name, columns=columns)
#         sheets.append(sheet)

#     return sheets


# @validate_arguments
# def format_sheet(
#     file: FilePath,
#     sheet_name: str,
#     *formats: ColumnFormat,
#     table_name: str="",
#     freeze_panes: "cell_refrence"="A2"
#     ) -> Path:
#     """Formats an existing sheet and optionally adds the sheet as a
#     Worksheet table with a default table format.

#     Provided formats for the 'patterns' argument use a Unix shell-style
#     wildcard as the key and the value is an XLSXFormat. The XLSXFormat is
#     typically created by using the column_format function.

#     Unix shell-style pattern
#     ╒═══════════╤══════════════════════════════════╕
#     │ Pattern   │ Meaning                          │
#     ╞═══════════╪══════════════════════════════════╡
#     │ *         │ matches everything               │
#     ├───────────┼──────────────────────────────────┤
#     │ ?         │ matches any single character     │
#     ├───────────┼──────────────────────────────────┤
#     │ [seq]     │ matches any character in seq     │
#     ├───────────┼──────────────────────────────────┤
#     │ [!seq]    │ matches any character not in seq │
#     ╘═══════════╧══════════════════════════════════╛
#     """
#     # Load file
#     wb = openpyxl.load_workbook(file)
#     ws = wb[sheet_name]

#     # Get column names and calc formats
#     xl_file = get_sheets(file_path, sheet_names=sheet_name)
#     rows = ws.rows
#     column_names = [cell.value for cell in next(rows)]
#     formats = Formats.from_user(file, sheet_name, *formats).to_dict()

#     # Loop by column and format column
#     for col in ws.iter_cols():
#         header = col[0]
#         cells = col[1:]
#         name = header.value
#         letter = header.column_letter

#         fmt = formats[name]

#         # Set width first so that text wrap works
#         ws.column_dimensions[letter].width = fmt.width

#         # Format header cell
#         header.alignment = fmt.header.alignment
#         header.number_format = fmt.header.number_format
#         header.font = fmt.header.font

#         # Add alignment and number format
#         for cell in cells:
#             cell.alignment = fmt.cell.alignment
#             cell.number_format = fmt.cell.number_format
#             cell.font = fmt.cell.font

#     # Freeze panes
#     ws.freeze_panes = ws[freeze_panes]

#     # Add table
#     table_name = table_name or sheet_name
#     table_name = table_name.replace(" ", "_")
#     table = Table(displayName=table_name, ref=ws.dimensions)
#     style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showColumnStripes=False, showRowStripes=True)
#     table.tableStyleInfo = style
#     ws.add_table(table)

#     # Save file
#     wb.save(str(file))
#     wb.close()

#     return file






# def format_column(ws: Worksheet, column: Column, format: Format):
#     cells = ws[column.letter]
#     header = cells[0]
#     data = cells[1:]

#     # Set width first so that text wrap works
#     ws.column_dimensions[column.letter].width = fmt.width

#     # Format header cell
#     header.alignment = fmt.header.alignment
#     header.number_format = fmt.header.number_format
#     header.font = fmt.header.font

#     # Add alignment and number format
#     for cell in data:
#         cell.alignment = fmt.cell.alignment
#         cell.number_format = fmt.cell.number_format
#         cell.font = fmt.cell.font

# # root = Path("~/py-analysis/appraisal-fees/appraisal_fees").expanduser()
# # file_path = root / "reports/Fee Reports.xlsx"
# # for sheet in get_sheets(file_path):
# #     print(f"{sheet.sheet_name=}")
# #     for column in sheet.columns:
# #         print(f"\t{column=}")


# class CellFormat(NamedTuple):
#     number_format: str
#     alignment: openpyxl.styles.Alignment
#     width: int
#     font: openpyxl.styles.Font


# class ColumnFormat(NamedTuple):
#     match: str
#     number_format: str=""
#     width: int=15
#     valign: str="top"
#     halign: str="left"
#     header_valign: str="bottom"
#     header_halign: str="left"
#     wrap_text: bool=True
#     num_align: bool=True
#     kwargs: Optional[Dict[str,Any]]=None

#     @property
#     def cell(self):
#         return CellFormat(
#             self.number_format,
#             openpyxl.styles.Alignment(
#                 vertical=self.valign,
#                 horizontal="right" if self.num_align and self.number_format else self.halign,
#                 wrap_text=self.wrap_text,
#             ),
#             self.width,
#             openpyxl.styles.Font(bold=False)
#         )

#     @property
#     def header(self):
#         return CellFormat(
#             self.number_format,
#             openpyxl.styles.Alignment(
#                 vertical=self.header_valign,
#                 horizontal="right" if self.num_align and self.number_format else self.header_halign,
#                 wrap_text=self.wrap_text,
#             ),
#             self.width,
#             openpyxl.styles.Font(bold=True)
#         )


# def get_cell_attrs(
#     file: Path,
#     sheet_name: str,
#     attr: str,
#     rows_to_review: int=100
#     ):
#     """Returns the dict of columns where each value is also a dict containing
#     the value counts of the attr values."""
#     wb = openpyxl.load_workbook(str(file), read_only=True)
#     ws = wb[sheet_name]
#     rows = ws.rows
#     header = next(rows)
#     # 1st row contains column headers
#     names = [cell.value for cell in header]
#     letters = [cell.column_letter for cell in header]
#     return tz.pipe(
#         # Loops over rows in sheet
#         rows,
#         tz.take(rows_to_review),
#         # For each row, create a dict usng names has keys
#         tz.map(lambda row: dict(zip(names, row))),
#         # Get the .xlsx data_type for each cell
#         tz.map(tz.valmap(tz.flip(getattr, attr))),
#         # Combine cells into a list per column
#         tz.merge_with(list),
#         # Count the cells for each data type in the column
#         tz.valmap(tz.frequencies),
#     )

# class Formats(NamedTuple):
#     """Aggregrates column formats for a given file and sheet_name."""
#     columns: Tuple[ColumnFormat]

#     @classmethod
#     def from_sheet(cls, file: Path, sheet_name: str, rows_to_review: int=100):
#         """Analyzes and existing sheet and returns a Patterns object which
#         provides default text alignment, column width, and number_format
#         options based upon how the .xlsx file has stored the data.
#         """
#         data_types = get_cell_attrs(file, sheet_name, "data_type",
#             rows_to_review)

#         return cls(tuple(
#             # If any cells have a date freq
#             ColumnFormat(column, "yyyy-mm-dd") if "d" in freq else
#             # If any cells have a string freq
#             ColumnFormat(column) if "s" in freq else
#             # Default data type in .xslx is "n" for number
#             ColumnFormat(column, "#,##0.00")
#             for column, freq in data_types.items()
#         ))


#     @classmethod
#     def from_user(
#         cls,
#         file: Path,
#         sheet_name: str,
#         *formats: ColumnFormat,
#         rows_to_review: int=100):
#         """Uses from_sheet to get default formats and then applies any formats
#         provided by the user."""

#         defaults = cls.from_sheet(file, sheet_name)
#         if len(formats) == 0:
#             return defaults

#         results = {}
#         for default in defaults.columns:
#             column = default.match
#             seen = column in results
#             for format in formats:
#                 assert isinstance(format, ColumnFormat), \
#                     "Provided format is not ColumnFormat"
#                 if fnmatch(column, format.match) and not seen:
#                     results[column] = format

#             if column not in results:
#                 results[column] = default

#         return cls(tuple(
#             v._replace(match=k)
#             for k, v in results.items()
#         ))

#     def to_dict(self) -> Dict[str,ColumnFormat]:
#         return {
#             column.match: column
#             for column in self.columns
#         }




# def to_xlsx(
#     data: Iterable[Dict],
#     file: Path,
#     sheet_name: str,
#     *formats: ColumnFormat,
#     apply_default_formats: bool=True,
#     over_write: bool=True
#     ) -> Path:
#     """Allows data to be appended to a .xlsx file as a new sheet.

#     If formats are provided, format_sheet will be called to format the sheet
#     and optionally adds the sheet as a Worksheet table with a default
#     table format.

#     Provided formats for the 'patterns' argument use a Unix shell-style
#     wildcard as the key and the value is an XLSXFormat. The XLSXFormat is
#     typically created by using the column_format function.

#     Unix shell-style pattern
#     ╒═══════════╤══════════════════════════════════╕
#     │ Pattern   │ Meaning                          │
#     ╞═══════════╪══════════════════════════════════╡
#     │ *         │ matches everything               │
#     ├───────────┼──────────────────────────────────┤
#     │ ?         │ matches any single character     │
#     ├───────────┼──────────────────────────────────┤
#     │ [seq]     │ matches any character in seq     │
#     ├───────────┼──────────────────────────────────┤
#     │ [!seq]    │ matches any character not in seq │
#     ╘═══════════╧══════════════════════════════════╛
#     """
#     if not over_write and file.exists():
#         wb = openpyxl.load_workbook(file)
#         ws = wb.create_sheet(sheet_name)
#     else:
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = sheet_name
#     for row_num, row in enumerate(data):
#         if row_num == 0:
#             ws.append([k for k in row])
#         ws.append(v for v in row.values())
#     wb.save(str(file))
#     if len(formats) > 0:
#         format_sheet(file, sheet_name, *formats)
#     elif apply_default_formats:
#         format_sheet(file, sheet_name)

#     wb.close()
#     return file


# append_sheet = partial(to_xlsx, over_write=False)
# append_sheet.__doc__ = "Same as to_xlsx except over_write=False"
